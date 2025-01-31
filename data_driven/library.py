import json
import jsonpath
import requests
import openpyxl


class Common:

    def __init__(self, file_name_path, sheet_name):
        global wk
        global sh
        wk = openpyxl.load_workbook(file_name_path)
        sh = wk[sheet_name]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range (1, c+1):
            cell = sh.cell(row=1, column=i)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self, row_number, jason_request, key_list):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row=row_number, column=i)
            jason_request[key_list[i-1]] = cell.value

        return jason_request




