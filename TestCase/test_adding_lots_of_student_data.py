import requests
import json
import jsonpath
import openpyxl


def test_adding_lots_of_student_data():
    api_url = 'https://thetestingworldapi.com/api/studentsDetails'
    file = open('C:\\Users\\dagle\\OneDrive\\Dokumenty\\API\\create_new_student.json')
    json_request = json.loads(file.read())

    wk = openpyxl.load_workbook('C:\\Users\\dagle\\OneDrive\\Dokumenty\\API\\table2.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row

    for i in range(2, rows+1):
        first_name_cell = sh.cell(rov=i, column=1)
        midd_name_cell = sh.cell(rov=i, column=2)
        last_name_cell = sh.cell(rov=i, column=3)
        data_cell = sh.cell(rov=i, column=4)

        json_request['first_name'] = first_name_cell.value
        json_request['midd_name_cell'] = midd_name_cell.value
        json_request['last_name_cell'] = last_name_cell.value
        json_request['data_cell'] = data_cell.value

        response = requests.post(api_url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201



